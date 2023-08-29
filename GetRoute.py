import json
import re
import requests
from openpyxl import load_workbook

def readJudge():
    with open('routeJugde.json', 'r',encoding='utf-8') as f:
        data = json.load(f)
    return data

def findFix(fix,ju):
    return ju['Fixes'].index(fix)


def proceedRoute(string: str):
    string = string.replace('、', ' ')
    string = string.replace('(', '')
    string = string.replace(')', '')
    string = string.replace('VOR', '')
    string = string.replace('NDB', '')
    string = string.replace('/', '')
    string = string.replace('。', '')
    string = string.replace('经', ' ')
    string = re.sub('[\u4e00-\u9fa5]', '', string)
    string = string.removeprefix(' ')
    spl = string.split(' ')
    for i in range(len(spl)):
        for n in range(len(spl[i]) - 1):
            if spl[i][n].isdigit() and (not spl[i][n + 1].isdigit()):
                spl[i] = spl[i][:n + 1] + ' ' + spl[i][n + 1:]
                break
    string = ''
    for i in range(len(spl)):
        string += spl[i]
        if i == len(spl) - 1:
            continue
        string += ' '
    return string


def judgeR(route,judge):

    route = route.split(' ')
    if len(route)<=2:
        return 0
    if route[1] not in judge:
        return 0
    pos1 = findFix(route[0], judge[route[1]])
    pos2 = findFix(route[2], judge[route[1]])

    j1 = pos1 - pos2
    if judge[route[1]]['WE'] == '\\':
        j2 = -1
    else:
        j2 = 1
    if j1 * j2 < 0:
        return 0
    else:
        return 1

def getRoute(icao):
    judge = readJudge()
    wb = load_workbook('./FLIGHT_AIRLINE.xlsx')
    sheet = wb['FLIGHT_AIRLINE']
    result = {icao: []}
    for i in range(sheet.max_row):
        start = sheet.cell(row=i + 1, column=4).value
        end = sheet.cell(row=i + 1, column=6).value
        route = sheet.cell(row=i + 1, column=10).value
        if start != icao or len(end) != 4:
            continue
        rou = proceedRoute(route)
        alt = judgeR(rou,judge)
        # if spl[1] == 'B330' or spl[1] == 'W233' or spl[1] == 'B213':
        #     alt = 0
        # elif spl[1] == 'G212' or spl[1] == 'W179':
        #     alt = 1
        # else:
        #     alt = -1
        newR = {'dep': start, 'arr': end, 'altType': alt, 'routes': rou}
        result[icao].append(newR)
    with open('ROUTE/' + icao + '.json', 'w') as f:
        json.dump(result, f)
    return result

def getAllRoute():
    judge = readJudge()
    wb = load_workbook('./FLIGHT_AIRLINE.xlsx')
    sheet = wb['FLIGHT_AIRLINE']
    result = {}
    for i in range(sheet.max_row):
        start = sheet.cell(row=i + 1, column=4).value
        end = sheet.cell(row=i + 1, column=6).value
        route = sheet.cell(row=i + 1, column=10).value
        rou = proceedRoute(route)
        spl = rou.split(' ')
        alt = judgeR(rou, judge)
        newR = {'dep': start, 'arr': end, 'routes': rou,'altType':alt}
        if start not in result:
            result[start] = {}
        if end not in result[start]:
            result[start][end] = []
        result[start][end].append(newR)
    with open('ROUTE/all.json', 'w') as f:
        json.dump(result, f)
    return result

# getRoute('VHHX')
getAllRoute()