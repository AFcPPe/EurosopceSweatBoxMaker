import json
import re

from openpyxl import load_workbook


route = {}
def convertCord(s:str):
    if s[0]=='E' or s[0]=='W':
        num = 4
    elif s[0]=='N' or s[0]=='S':
        num =3
    else:
        return ''
    if s[0] == 'N' or s[0] == 'E':
        prefix = 1
    elif s[0] == 'W' or s[0] == 'S':
        prefix = -1
    return (int(s[1:num])+(int(s[num+2:])*10/600 +int(s[num:num+2]))*10/600)*prefix

def getRouteId():
    wb = load_workbook('./EN_ROUTE_RTE.xlsx')
    sheet = wb['EN_ROUTE_RTE']
    for i in range(2,sheet.max_row+1):
        route[sheet.cell(row=i, column=1).value] = {'name':sheet.cell(row=i, column=2).value,'Fixes':[]}

def getWE():
    curr = ''
    wb = load_workbook('./RTE_SEG.xlsx')
    sheet = wb['RTE_SEG']
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=2).value!=curr:
            curr = sheet.cell(row=i, column=2).value
            route[curr]['start'] = [convertCord(sheet.cell(row=i, column=10).value),convertCord(sheet.cell(row=i, column=9).value)]
            route[curr]['Fixes'].append(sheet.cell(row=i, column=6).value)
        route[curr]['end'] = [convertCord(sheet.cell(row=i, column=19).value),
                              convertCord(sheet.cell(row=i, column=18).value)]
        route[curr]['Fixes'].append(sheet.cell(row=i, column=15).value)

    for each in route:
        wc = route[each]['start'][0]-route[each]['end'][0]
        jc = route[each]['start'][1]-route[each]['end'][1]
        if wc*jc<0:
            route[each]['WE'] = '\\'

        else:
            route[each]['WE'] = '/'
        if wc>0:
            route[each]['Fixes'] = list(reversed(route[each]['Fixes']))


getRouteId()
getWE()
with open('routeJugde.json','w')as f:
    json.dump(route,f)

